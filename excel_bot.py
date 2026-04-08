# excel_bot.py
# pip install pandas openpyxl

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os

INPUT_CSV = 'sales_data.csv'
OUTPUT_XLSX = f'sales_report_{datetime.today().strftime("%Y%m%d")}.xlsx'

def fill(hex_color):
    return PatternFill('solid', fgColor=hex_color)

def bold(size=10, color='000000'):
    return Font(bold=True, size=size, color=color)

def thin_border():
    s = Side(style='thin', color='E2E8F0')
    return Border(left=s, right=s, top=s, bottom=s)

# ---- Load & analyze data ----------------------------------
def load_data(path):
    df = pd.read_csv(path)
    df['Revenue'] = pd.to_numeric(df['Revenue'], errors='coerce').fillna(0)
    return df

def metrics(df):
    return {
        'total': round(df['Revenue'].sum(), 2),
        'avg': round(df['Revenue'].mean(), 2),
        'top_prod': df.groupby('Product')['Revenue'].sum().idxmax(),
        'top_sp': df.groupby('Salesperson')['Revenue'].sum().idxmax(),
        'count': len(df),
        'units': df['Units Sold'].sum()   # ✅ FIX HERE
    }
# ---- Build styled data sheet ------------------------------
def build_data_sheet(ws, df):
    ws.title = 'Sales Data'
    headers = list(df.columns)

    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = fill('0F1923')
        c.font = bold(10, 'FFFFFF')
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = thin_border()

    ws.row_dimensions[1].height = 22

    for ri, row in enumerate(df.itertuples(index=False), 2):
        alt = (ri % 2 == 0)

        for ci, val in enumerate(row, 1):
            c = ws.cell(row=ri, column=ci, value=val)
            c.fill = fill('F1F5F9') if alt else fill('FFFFFF')
            c.border = thin_border()
            c.alignment = Alignment(horizontal='center', vertical='center')

            if ci == len(headers):  # Revenue column
                c.number_format = '"$"#,##0.00'
                if isinstance(val, (int, float)) and val >= 5000:
                    c.fill = fill('F59E0B')
                    c.font = bold(10)

        ws.row_dimensions[ri].height = 18

    # Auto-size
    for ci, h in enumerate(headers, 1):
        vals = [len(str(ws.cell(r, ci).value or '')) for r in range(2, ws.max_row + 1)]
        ws.column_dimensions[get_column_letter(ci)].width = min(max(len(h), max(vals, default=0)) + 4, 30)

    ws.freeze_panes = 'A2'

# ---- Build dashboard summary sheet -----------------------
def build_dashboard(ws, m, df):
    ws.title = 'Dashboard'

    ws.merge_cells('A1:D1')
    t = ws['A1']
    t.value = 'Sales Report Dashboard'
    t.fill = fill('0F1923')
    t.font = bold(14, 'FFFFFF')
    t.alignment = Alignment(horizontal='center', vertical='center')

    ws.row_dimensions[1].height = 30

    kpis = [
        ('Total Revenue', f'${m["total"]:,.2f}', '0D9488'),
        ('Avg Sale Value', f'${m["avg"]:,.2f}', '0D9488'),
        ('Total Units', str(m['units']), '3B82F6'),
        ('Top Product', m['top_prod'], 'F59E0B'),
        ('Top Salesperson', m['top_sp'], '22C55E'),
        ('Records', str(m['count']), '64748B'),
    ]

    for i, (label, value, color) in enumerate(kpis, 3):
        lc = ws.cell(row=i, column=1, value=label)
        lc.fill = fill(color)
        lc.font = bold(10, 'FFFFFF')
        lc.alignment = Alignment(horizontal='left', vertical='center', indent=1)

        ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
        vc = ws.cell(row=i, column=2, value=value)
        vc.font = bold(11)
        vc.alignment = Alignment(horizontal='center', vertical='center')

        ws.row_dimensions[i].height = 24

    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 20

# ---- Main ------------------------------------------------
def run():
    print('Excel Bot starting...')

    df = load_data(INPUT_CSV)
    m = metrics(df)

    print(f'Loaded {m["count"]} records | Revenue: ${m["total"]:,.2f}')

    wb = Workbook()
    ws_data = wb.active
    ws_dash = wb.create_sheet()

    build_data_sheet(ws_data, df)
    build_dashboard(ws_dash, m, df)

    wb.move_sheet(ws_dash, offset=-1)

    wb.save(OUTPUT_XLSX)

    print(f'Saved: {OUTPUT_XLSX}')
    print(f'Path : {os.path.abspath(OUTPUT_XLSX)}')


if __name__ == '__main__':
    run()